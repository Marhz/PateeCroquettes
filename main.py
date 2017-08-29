import openpyxl
from datetime import datetime
import sys

def getDate(filename):
    date = filename[0:6]
    date = datetime.strptime(date, "%d%m%y")
    return date

def getSide(filename):
    return "A" if "gauche" in filename else "B"

def addFoodType(file, dates):
    summary.cell(row=1, column=7).value = "Type"
    for rowNum in range(2, summary.max_row + 1):
        summary.cell(row=rowNum, column=7).value = "LOL"

formatDate = lambda d: datetime.strptime(d, '%d/%m/%Y %H:%M:%S')

filename = sys.argv[1]
date = getDate(filename)
side = getSide(filename)
print("Loading file...")
doc = openpyxl.load_workbook(filename)
print("OK!")
fillSheet = doc.get_sheet_by_name("Bowl_fill_history")
fillSheet = [r for r in fillSheet.rows if r[3].value[0] == side]
formattedRow = {}
begEnd = {}
print("Mapping dates...")
for row in fillSheet:
    if row[1].value in formattedRow.keys():
        formattedRow[row[1].value].append(row[4].value)
    else:
        formattedRow[row[1].value] = [row[4].value]
for key in formattedRow.keys():
    for item in formattedRow[key]:
        item = formatDate(item)
        if item.day == date.day and item.hour == 8 and item.minute > 45 and item.minute < 59:
            if key in begEnd.keys():
                begEnd[key]['begin'] = item
            else:
                begEnd[key] = {'begin': item}
    if "begin" not in begEnd[key].keys():
        begEnd[key]['begin'] = formatDate(str(date.day) + "/" + str(date.moth) + "/" + str(date.year) + " 08:50:00")
    for item in formattedRow[key]:
        item = formatDate(item)
        if item.day == date.day and item.hour == 10 and item.minute < 30:
            begEnd[key]['end'] = item
            break
    if "end" not in begEnd[key].keys():
        begEnd[key]['end'] = formatDate(str(date.day)+"/"+str(date.month)+"/"+str(date.year)+" 10:05:00")
print("OK!")
print("Writing values...")
summary = doc.get_sheet_by_name("Summary_with_steelyard")
addFoodType(summary, begEnd)
doc.save(filename)
print("All done!")